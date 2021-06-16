from math import pi
import collections
import random
import uuid
import json
import openpyxl
from openpyxl import Workbook


# 打印九九乘法表
for i in range(1, 10):
    for j in range(1, i+1):
        print(j, '*', i, '=', i*j, end='  ')
    else:
        print(' ')


# 两数之和
nums = [2, 7, 11, 15]
target = 9
for i in range(0, len(nums)):
    for j in range(i + 1, len(nums)):
        if nums[i] + nums[j] == target:
            print([i, j])

nums = [2, 7, 3, 6]
target = 9
result = []
for i in range(0, len(nums)):
    for j in range(i + 1, len(nums)):
        if nums[i] + nums[j] == target:
            result.append([i, j])
            continue
print(result)

# 20210428
# 1. 请设计一个函数，函数的输入是数字，打印这个数是奇数还是偶数


def odd_or_even(number):
    if number % 2 == 0:
        return '偶数'
    else:
        return '奇数'


print(odd_or_even(5))

# 2. 请设计一个函数，输入是半径，返回值是对应圆的面积


def area_circle(r):
    return pi * (r ** 2)


print(area_circle(2))

# 3. 给到一个列表，请计算列表中所有元素之和
nums = [1, 4, 6, 9, 3.5]
result = 0
for i in range(0, len(nums)):
    result = result + nums[i]

print(result)

# 4. 给到一个列表，请计算出列表中所有偶数之和
nums = [1, 4, 6, 9, 3.5]
sum_even = 0
for i in range(0, len(nums)):
    if nums[i] % 2 == 0:
        sum_even = sum_even + nums[i]

print(sum_even)

# 5. 给到一个列表，请计算出列表中所有素数之和
nums = [1, 2, 11, 101, 4, 6, 9]
sum_prime = 0
for i in range(0, len(nums)):
    if nums[i] == 1:
        continue
    for j in range(2, nums[i]):
        if nums[i] % j == 0:
            break
    else:
        sum_prime = sum_prime + nums[i]

print(sum_prime)

# 6. 将 3，4，5 封装成函数，函数的输入是列表，函数的返回值为对应的和
nums = [1, 2, 11, 101, 4, 6, 9]


def sum_list(nums):
    result = 0
    for i in range(0, len(nums)):
        result = result + nums[i]
    if i == len(nums) - 1:
        return result


print(sum_list(nums))


def sum_list(nums):
    result = 0
    for i in range(0, len(nums)):
        result = result + nums[i]
        if i == len(nums) - 1:
            return result


print(sum_list(nums))


def sum_even(nums):
    result = 0
    for i in range(0, len(nums)):
        if nums[i] % 2 == 0:
            result = result + nums[i]
    if i == len(nums) - 1:
        return result


print(sum_even(nums))


def sum_prime(nums):
    result = 0
    for i in range(0, len(nums)):
        if nums[i] == 1:
            continue
        for j in range(2, nums[i]):
            if nums[i] % j == 0:
                break
        else:
            result = result + nums[i]
    if i == len(nums) - 1:
        return result


print(sum_prime(nums))

# 7. 请打印出 1-100 之间所有的素数
for i in range(1, 101):
    if i == 1:
        continue
    for j in range(2, i):
        if i % j == 0:
            break
    else:
        print(i)

for i in range(1, 101):
    if i == 1:
        continue
    for j in range(2, i//2+1):
        if i % j == 0:
            break
    else:
        print(i)

# 8. 打印指定范围的素数
a = 9
b = 50
for i in range(a, b):
    if i == 1:
        continue
    for j in range(2, i):
        if i % j == 0:
            break
    else:
        print(i)

# 9. 请设计一个函数，输入是一个范围，返回这个范围内的所有素数


def range_prime(a, b):
    nums = []
    for i in range(a, b+1):
        if i == 1:
            continue
        for j in range(2, i):
            if i % j == 0:
                break
        else:
            nums.append(i)
        if i == b:
            return nums


a = 9
b = 50
print(range_prime(a, b))

# 20210430
# 1. 请使用冒泡排序对列表中的元素进行排序
arr = [22, 85, 19, 98, 8, 24, 5, 23, 3, 11, 2]
for i in range(len(arr) - 1):  # 设置冒泡排序次数
    for j in range(0, len(arr) - i - 1):  # 设置每一次冒泡排序需要的排序次数
        if arr[j] > arr[j + 1]:
            arr[j], arr[j + 1] = arr[j + 1], arr[j]  # 将较小的数排在前面

print(arr)

# 2. 请使用选择排序对列表中的元素进行排序
# 每一次遍历的时候，一选到小的就交换，直到遍历结束i位置上就是所有的最小值
arr = [22, 85, 19, 98, 8, 24, 5, 23, 3, 11, 2]
for i in range(0, len(arr)):  # 设置选择排序次数
    for j in range(i + 1, len(arr)):  # i之后的所有数与i比较大小
        min_index = i  # 迭代得出最小值的索引(j每增加一次就在与前面的最小值对比)
        if arr[j] < arr[min_index]:
            min_index = j
        arr[i], arr[min_index] = arr[min_index], arr[i]  # 每比较一次就将最小值放在i位置上

print(arr)

arr = [22, 85, 19, 98, 8, 24, 5, 23, 3, 11, 2]
for i in range(0, len(arr)):  # 设置选择排序次数
    for j in range(i + 1, len(arr)):  # i之后的所有数与i比较大小
        if arr[j] < arr[i]:
            arr[i], arr[j] = arr[j], arr[i]  # 每比较一次就将最小值放在i位置上

print(arr)

# 每一次遍历后选出最小值与i交换位置
arr = [22, 85, 19, 98, 8, 24, 5, 23, 3, 11, 2]
for i in range(0, len(arr)):  # 设置选择排序次数
    min_index = i
    for j in range(i + 1, len(arr)):  # i之后的所有数与i比较大小
        if arr[j] < arr[min_index]:
            min_index = j
    arr[i], arr[min_index] = arr[min_index], arr[i]

print(arr)


# 20210501
# 1.请设计一个函数，查找列表中的最小的元素
def min_factor(arr):
    for i in range(1, len(arr)):
        if arr[i] < arr[0]:
            arr[0], arr[i] = arr[i], arr[0]
    return arr[0]


arr = [22, 85, 19, 98, 8, 24, 5, 23, 3, 11, 2]
print(min_factor(arr))


# 2.现在有一个字典，key的类型为字符串，value类型为int，请设计一个函数，求所有的value之和
# 字典
def sum_value(dictionary):
    result = 0
    for i in dictionary:  # 重点
        result = result + dictionary[i]
    return result


dictionary = {'xsw': 10, 'lyj': 60, 'zd': 120}
print(sum_value(dictionary))


# 列表
def sum_value(arr):
    result = 0
    for i in range(len(arr)):  # 与字典的区别
        result = result + arr[i]
    return result


arr = [10, 60, 120]
print(sum_value(arr))


# 3.现在有一个排好序的列表，请设计一个函数:
#   函数的输入是一个元素
#   函数内部使用二分查找查找该元素
#   如果查找成功，则返回该元素对应的索引
#   如果列表中不存在该元素，则返回None


def binary_search(arr, l, r, x):  # arr为排好序的列表，l为列表最左的下标，r为最右的下标，x为要查找的元素
    if r >= l:
        mid = l + (r - l) // 2
        if x == arr[mid]:
            return mid
        if x < arr[mid]:
            return binary_search(arr, l, mid - 1, x)  # 递归
        if x > arr[mid]:
            return binary_search(arr, mid + 1, r, x)  # 递归
        else:
            return None


arr = [1, 3, 6, 8, 10, 15, 17, 20]
print(binary_search(arr, 0, len(arr) - 1, 15))
print(binary_search(arr, 0, len(arr) - 1, 2))

# 4.现在给到一个列表，请统计列表中各元素的个数
arr = [2, 5, 19, 8, 5, 24, 5, 2, 3, 11, 2]
c = collections.Counter(arr)  # c.subtract可以从计数结果中减去个数，c.update可以增加元素统计
print(c)
dict_c = dict(c)
print(dict_c)

result = dict()
for i in arr:
    result[i] = arr.count(i)

print(result)


# 20210507
# 1. 现在有一张表，表shema如下，请使用 python 为该表随机生成5000行数据，只需将 insert 语句打印到控制台。
#       其中，id 需要唯一(使用uuid),其他数据请随机值填充，创建时间和更新时间请使用 NULL 填充
#       CREATE TABLE `fact_store_info`  (
#          `id` int NOT NULL,
#          `name` varchar(255) NULL COMMENT '门店名字',
#          `store_no` varchar(255) NULL COMMENT '门店编码',
#          `province` varchar(255) NULL COMMENT '所属省份',
#          `city` varchar(255) NULL COMMENT '所属城市',
#          `channel` varchar(255) NULL COMMENT '所属渠道',
#          `created_time` timestamp NULL COMMENT '创建时间',
#          `updated_time` timestamp NULL COMMENT '更新时间',
#          PRIMARY KEY (`id`)
#        );
#        tips:
#            a. 随机数据
fact_store_info = {'id': str(uuid.uuid1()),
                   'name': random.randint(1, 100),
                   'store_no': random.randint(1, 100),
                   'province': random.randint(1, 100),
                   'city': random.randint(1, 100),
                   'channel': random.randint(1, 100),
                   'created_time': 'NULL',
                   'updated_time': 'NULL'}
print(fact_store_info)
print('''
insert into fact_store_info (id,name,store_no,province,city,channel,created_time,updated_time) \
values({0[id]:s}, {0[name]:d}, {0[store_no]:d}, {0[province]:d}, {0[city]:d}, {0[channel]:d}, \
{0[created_time]:s}, {0[updated_time]:s});
'''.format(fact_store_info))

for i in range(5):
    fact_store_info = {'id': str(uuid.uuid1()),
                       'name': random.randint(1, 100),
                       'store_no': random.randint(1, 100),
                       'province': random.randint(1, 100),
                       'city': random.randint(1, 100),
                       'channel': random.randint(1, 100),
                       'created_time': 'NULL',
                       'updated_time': 'NULL'}
    print('''insert into fact_store_info (id,name,store_no,province,city,channel,created_time,updated_time) \
    values({id:s}, {name:d}, {store_no:d}, {province:d}, {city:d}, {channel:d}, {created_time:s}, {updated_time:s}); \
    '''.format(**fact_store_info))

# 2. 请将 1 中的 insert 语句输出到文件中，文件名为 fact_store_info.sql


def generate_table_data(file_name, rows):
    with open(file_name, 'w+') as file:
        for i in range(rows):
            id = str(uuid.uuid1())
            name = str(random.randint(1, 100))
            store_no = str(random.randint(1, 100))
            province = str(random.randint(1, 100))
            city = str(random.randint(1, 100))
            channel = str(random.randint(1, 10))
            created_time = 'NULL'
            updated_time = 'NULL'
            sql = "insert into fact_store_info (id,name,store_no,province,city,channel,created_time,updated_time)" \
                  "values({0}, {1}, {2}, {3}, {4}, {5}, {6}, {7});" \
                .format(id, name, store_no, province, city, channel, created_time, updated_time)
            file.write(sql)
            file.write('\n')

# generate_table_data("/Users/xiaoshuwen/PycharmProjects/learn_python/python_basis/fact_store_info.sql", 5000)


# 20210508
# 现在有一个列表，请使用插入排序对该列表进行排序
# 冒泡排序


arr = [5, 18, 23, 9, 64, 2, 0, 37, 15]
for i in range(1, len(arr)):
    for j in range(0, i):
        if arr[i] < arr[j]:
            arr[j], arr[i] = arr[i], arr[j]

print(arr)

# 插入排序
arr = [5, 18, 23, 9, 64, 2, 0, 37, 15]
for i in range(1, len(arr)):
    j = i - 1
    key = arr[i]
    # 将arr[i]与它前面的每一个数比较大小，如果arr[i]更小，则前面的数就往后挪一位，直到arr[i]到达合适位置，或者到第一位
    while j >= 0 and key < arr[j]:
        arr[j + 1] = arr[j]
        j = j - 1
    arr[j + 1] = key  # j<0或者arr[j]>arr[i]时,循环停止，arr[i]换到j+1的位置上

print(arr)


# 20210509
# 1. 请设计一个车类，属性有车的颜色，价格，行为(方法)有运输(transport)、启动(start)、停止(stop)
class Car:
    def __init__(self, color, price):
        self.color = color
        self.price = price

    @staticmethod
    def transport():
        print("运输")

    @staticmethod
    def start():
        print("启动")

    @staticmethod
    def stop():
        print("停止")


mycar = Car('white', '10w')
print(mycar)
print(mycar.color)
print(mycar.price)
mycar.transport()
mycar.start()
mycar.stop()

# 2. 请设计一个宝马车类，继承车类


class BMW(Car):
    pass

# 笔试题：B为一个字符串，A是B的子集（字母顺序不能变），求A的可能结果个数。注：len(A)<=len(B),否则结果为0；B中有重复元素，结果为0。


# 20210510
# 给你两个有序整数数组 nums1 和 nums2，请你将 nums2 合并到 nums1 中，使 nums1 成为一个有序数组。
# 初始化 nums1 和 nums2 的元素数量分别为 m 和 n 。你可以假设 nums1 的空间大小等于 m + n，这样它就有足够的空间保存来自 nums2 的元素。
# 输入：nums1 = [1,2,3,0,0,0], m = 3, nums2 = [2,5,6], n = 3
# 输出：[1,2,2,3,5,6]
# https://leetcode-cn.com/problems/merge-sorted-array/solution/python3-shuang-zhi-zhen-xiu-gai-0mnfu-za-dee5/

nums1 = [1, 3, 5, 7]
nums2 = [2, 3, 4, 8, 10]

# 插入排序，短数组插入长数组里面去
if len(nums2) <= len(nums1):
    nums = nums1 + nums2
    for i in range(len(nums1), len(nums)):
        j = i - 1
        key = nums[i]
        while j >= 0 and key < nums[j]:
            nums[j+1] = nums[j]
            j = j - 1
        nums[j + 1] = key
else:
    nums = nums2 + nums1
    for i in range(len(nums2), len(nums)):
        j = i - 1
        key = nums[i]
        while j >= 0 and key < nums[j]:
            nums[j+1] = nums[j]
            j = j - 1
        nums[j + 1] = key

print(nums)


# 20210512
# 1. 请用自己的话描述数据立方体
# cube生成多维数据集，结果集包含各维度的所有可能组合
# 语法；GROUP BY 列1，列2，... WITH CUBE
# 例：GROUP BY country,city WITH CUBE 等效于 GROUP BY GROUPING SETS ((country,city),(country),(city),())
# 在group by子句中有n个列或者是有n个表达式的话，结果集会返回2的n次幂个可能组合

# 可以使用 GROUPING 函数区分 CUBE 操作生成的 NULL 值和在实际数据中返回的 NULL 值
# 如果列值来自事实数据，GROUPING 函数将返回 0；如果列值是由 CUBE 操作生成的 NULL，则返回 1
# 在 CUBE 操作中，生成的 NULL 代表所有值。
# 可以编写 SELECT 语句以使用 GROUPING 函数将生成的任一 NULL 替换为字符串 ALL。
# 由于事实数据中的 NULL 表示数据值未知，因此也可以将 SELECT 编码为返回字符串 UNKNOWN，用于表示事实数据中的 NULL
# SELECT CASE WHEN (GROUPING(Item) = 1) THEN 'ALL'
#             ELSE ISNULL(Item, 'UNKNOWN')
#        END AS Item,
#        CASE WHEN (GROUPING(Color) = 1) THEN 'ALL'
#             ELSE ISNULL(Color, 'UNKNOWN')
#        END AS Color,
#        SUM(Quantity) AS QtySum
# FROM Inventory
# GROUP BY Item, Color WITH CUBE

# 2. 项目 data 目录下有一个 area_code_2021.json 文件，该文件为省市数据，
#    请解析该文件，从该文件中提取所有的省份，然后写到 data 目录下的province.txt 文件中
def json_to_txt(json_file, txt_file):
    with open(json_file, 'r') as file:
        arr = json.load(file)
        with open(txt_file, 'a+') as province_file:
            for i in range(len(arr)):
                province = arr[i]['name']
                province_file.write(province)
                province_file.write('\n')


# json_to_txt('/Users/xiaoshuwen/PycharmProjects/learn_python/data/area_code_2021.json',
#             '/Users/xiaoshuwen/PycharmProjects/learn_python/data/province.txt')


# 3. 请设计一个 calc 模块，该模块提供基本的数学运算


class Calc(object):
    def __init__(self, num1, num2):
        self.num1 = num1
        self.num2 = num2

    # 加法
    def add(self):
        return self.num1 + self.num2

    # 减法
    def subtract(self):
        return self.num1 - self.num2

    # 乘法
    def multiply(self):
        return self.num1 * self.num2

    # 除法
    def divide(self):
        return self.num1 / self.num2

    # 取余
    def mod(self):
        return self.num1 % self.num2

    # 取绝对值
    def abs(self):
        if self.num1 >= 0:
            if self.num2 >= 0:
                return self.num1, self.num2
            else:
                return self.num1, -self.num2
        if self.num1 < 0:
            if self.num2 >= 0:
                return -self.num1, self.num2
            else:
                return -self.num1, -self.num2


a = -13
b = 4
test = Calc(a, b)
print(test.add())
print(test.subtract())
print(test.multiply())
print(test.divide())
print(test.mod())
print(test.abs())


# 20210514
# 1. 提取 area_code_2021.json 省份信息，要求包含省份代码(code),省份名，
#    并将这些信息写入到文件名为 area.xlsx sheet 名为 province 中
def json_to_xlsx(json_file, xlsx_file):
    with open(json_file, 'r') as file:
        arr = json.load(file)
        with open(xlsx_file, 'a+') as province_file:
            wb = Workbook()  # 获取工作簿/Excel文件,同时也至少创建了一个工作表/worksheet
            ws = wb.active  # 调用当前工作表
            ws.title = "province"
            for i in range(len(arr)):
                code = str(arr[i]['code'])
                province = arr[i]['name']
                ws.cell(i+1, 1, code)
                ws.cell(i+1, 2, province)
                # ws['A' + str(i+1)] = code
                # ws['B' + str(i+1)] = province
            wb.save(xlsx_file)


json_to_xlsx('/Users/xiaoshuwen/PycharmProjects/learn_python/data/area_code_2021.json',
             '/Users/xiaoshuwen/PycharmProjects/learn_python/data/area.xlsx')

# 2. 提取 area_code_2021.json 省份城市信息，要求包含城市代码(code),省份名，城市名，
#    并将这些信息写入到文件名为 area.xlsx sheet 名为 city 中


# def json_to_xlsx(json_file, xlsx_file):
#     with open(json_file, 'r') as file:
#         arr = json.load(file)
#         with open(xlsx_file, 'a+') as city_file:
#             wb = Workbook()
#             # ws1 = wb.active
#             # ws1.title = "province"
#             ws2 = wb.create_sheet("city")
#             for i in range(len(arr)):
#                 province = arr[i]['name']
#                 for j in range(len(arr[i]['children'])):
#                     code = str(arr[i]['children'][j]['code'])
#                     city = arr[i]['children'][j]['name']
#                     city_file.write(code)
#                     city_file.write('\t')
#                     city_file.write(province)
#                     city_file.write('\t')
#                     city_file.write(city)
#                     city_file.write('\n')


def json_to_xlsx(json_file, xlsx_file):
    with open(json_file, 'r') as file:
        arr = json.load(file)
        with open(xlsx_file, 'a+') as city_file:
            wb = openpyxl.load_workbook(xlsx_file)
            ws2 = wb.create_sheet("city")
            count = 0
            for i in range(len(arr)):
                province = arr[i]['name']
                for j in range(len(arr[i]['children'])):
                    code = str(arr[i]['children'][j]['code'])
                    city = arr[i]['children'][j]['name']
                    count += 1
                    ws2['A' + str(count)] = code
                    ws2['B' + str(count)] = province
                    ws2['C' + str(count)] = city
            wb.save(xlsx_file)


json_to_xlsx('/Users/xiaoshuwen/PycharmProjects/learn_python/data/area_code_2021.json',
             '/Users/xiaoshuwen/PycharmProjects/learn_python/data/area.xlsx')

# 3. 最大子序列和，给定一个整数列表 nums ，找到一个具有最大和的连续子数组（子数组最少包含一个元素），返回其最大和。
#    eg: nums = [-2,1,-3,4,-1,2,1,-5,4] 输出为 6
#    tips: 连续子数组 [4,-1,2,1] 的和最大，为 6
nums = [-2, 1, -3, 4, -1, 2, 1, -5, 4]
max_sum = nums[0]
for i in range(len(nums)):
    sum = nums[i]
    for j in range(i+1, len(nums)):
        sum += nums[j]
        if sum > max_sum:
            max_sum = sum

print(max_sum)

# 20210518
# 实现 pow(x, n) ，即计算 x 的 n 次幂函数


def pow(x, n):
    if x == 0:
        return 0
    if n == 0:
        return 0
    if n > 0:
        result = 1
        for i in range(n):
            result = result * x
        return result
    if n < 0:
        result = 1
        for i in range(-n):
            result = result * x
        return 1/result


print(pow(2, 3))
print(pow(0, 4))
print(pow(-2, 3))
print(pow(-2, -3))
print(pow(-1/2, -3))


# 20210519
# https://leetcode-cn.com/problems/simplify-path/
def simplify_path(path):
    paths = path.split('/')
    result = []
    for ele in paths:
        # ..result非空则返回上一级文件名
        if ele == '..' and result:
            result.pop()
        # .或者空字符则跳过，不是则加入result
        if ele not in ('.', '', '..'):
            result.append(ele)
    # 用'/'拼接result，并且在最前面加上'/'
    return '/'+'/'.join(result)


path1 = "/home//foo/"
path2 = "/a/./b/../../c/"
path3 = "/../"
print(simplify_path(path2))


# 20210520
# https://leetcode-cn.com/problems/subsets/
# 给你一个整数数组 nums ，数组中的元素 互不相同 。返回该数组所有可能的子集（幂集）。
# 解集 不能 包含重复的子集。你可以按 任意顺序 返回解集。
# https://www.runoob.com/w3cnote/python-understanding-dict-copy-shallow-or-deep.html 直接赋值、浅拷贝和深度拷贝解析


def subsets(nums):
    res = []

    def backtracking(nums, idx, path):
        res.append(path.copy())
        for i in range(idx, len(nums)):
            path.append(nums[i])
            backtracking(nums, i + 1, path)
            path.pop()

    backtracking(sorted(nums), 0, [])
    return res


nums = [1, 2, 3]
print(subsets(nums))






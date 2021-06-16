#! python
# Doc
# 该脚本解析crm.store_additional表的excel模版文件，并生成sql
# python xxx.py 参数1 参数2 参数3
# 参数1: excel 绝对路径
# 参数2: sheet 名
# 参数3: 结果文件名，如果没有提供，则默认为 result.sql，文件路径为当前shell路径

import sys
import os
from openpyxl import load_workbook


def parse_excel(filename, sheetname, result_file, database, table):
    wb = load_workbook(filename)
    ws = wb[sheetname]

    sql_values = []
    delete_store_codes = []

    # 读数据,从第二行开始
    for row_index in range(2, ws.max_row + 1):
        store_type = ws[str('A' + str(row_index))].value
        cn = ws[str("B" + str(row_index))].value
        region = ws[str('C' + str(row_index))].value
        city = ws[str('D' + str(row_index))].value
        type = ws[str('E' + str(row_index))].value
        store_name_zh = ws[str('F' + str(row_index))].value
        store_name_en = ws[str('G' + str(row_index))].value
        store_address_zh = ws[str('H' + str(row_index))].value
        store_address_en = ws[str('I' + str(row_index))].value
        store_phone = ws[str('J' + str(row_index))].value
        store_manager = ws[str('K' + str(row_index))].value
        store_manager_phone = ws[str('L' + str(row_index))].value
        store_area_manager = ws[str('M' + str(row_index))].value
        store_area_manager_phone = ws[str('N' + str(row_index))].value
        store_region_manager = ws[str('O' + str(row_index))].value
        store_region_manager_phone = ws[str('P' + str(row_index))].value
        start_date = ws[str('Q' + str(row_index))].value
        email = ws[str('R' + str(row_index))].value
        head = ws[str('S' + str(row_index))].value

        tb_store_code = cn.upper()
        delete_store_codes.append(tb_store_code)

        if store_type.startswith('OR'):
            tb_store_classification = "OR"
            tb_store_type = type
        else:
            tb_store_classification = "FR"
            tb_store_type = 'Store'

        tb_store_name_display = store_name_en if store_name_zh is None or store_name_zh == "" else store_name_zh
        tb_store_name = tb_store_name_display if tb_store_classification == 'OR' else store_type + tb_store_name_display

        tb_store_type_local = type
        tb_region = region
        tb_city = city
        tb_director = head
        tb_rom = store_region_manager
        tb_dom = store_area_manager

        sql_value = "('{0}','{1}','{2}','{3}','{4}',{5},{6},{7},NULL,{8},NULL,{9},{10})\n" \
            .format(tb_store_code,
                    tb_store_name,
                    tb_store_name_display,
                    tb_store_classification,
                    tb_store_type,
                    'NULL' if tb_store_type_local is None else "\'{}\'".format(tb_store_type_local),
                    'NULL' if tb_region is None else "\'{}\'".format(tb_region),
                    'NULL' if tb_city is None else "\'{}\'".format(tb_city),
                    'NULL' if tb_director is None else "\'{}\'".format(tb_director),
                    'NULL' if tb_rom is None else "\'{}\'".format(tb_rom),
                    'NULL' if tb_dom is None else "\'{}\'".format(tb_dom))
        sql_values.append(sql_value)

    with open(result_file, 'w') as file:
        delete_codes = ""
        for store_code in delete_store_codes:
            delete_codes += ",\'{}\'".format(store_code)
        delete_sql_prefix = "DELETE FROM {}.{} \n" \
                            "WHERE store_code IN (".format(database, table)
        delete_sql = delete_sql_prefix + delete_codes[1:] + ")\n;"
        file.write(delete_sql)
        insert_sql_prefix = "INSERT INTO {}.{}" \
                            "(store_code,store_name,store_name_display,store_classification,store_type,store_type_local,region,city,comp,director,store_state,rom,dom)\n" \
                            "VALUES".format(database, table)
        for sql_value in sql_values:
            insert_sql = insert_sql_prefix + sql_value + ";"
            file.write("\n")
            file.write(insert_sql)


if __name__ == "__main__":
    args = sys.argv
    pwd = os.getcwd()
    if len(args) <= 2:
        exit(-1)
    filename = args[1]
    sheetname = args[2]
    result_file = pwd + "/" + args[3] if len(args) == 4 else pwd + "/result.sql"
    default_db = "ad_tmp_test"
    default_table = "store_additional"
    parse_excel(filename, sheetname, result_file, default_db, default_table)
